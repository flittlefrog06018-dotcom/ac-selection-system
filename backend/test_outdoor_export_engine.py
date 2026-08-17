import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
import os
import json
import sys

sys.stdout.reconfigure(encoding='utf-8')

def test_outdoor_db_and_export():
    excel_db_path = "backend/product_database/EQUIPMENT_Data.xlsx"
    template_path = "backend/product_database/選機表-.xlsx"
    
    wb_db = openpyxl.load_workbook(excel_db_path, data_only=True)
    ws_outdoor = wb_db["outdoor_units"]
    
    cols = list(ws_outdoor.columns)
    outdoor_units = []
    for c in range(2, len(cols)):
        sys_val = cols[c][1].value
        ser_val = cols[c][2].value
        mod_val = cols[c][3].value
        cap_val = cols[c][4].value
        nom_val = cols[c][5].value
        pwr_sup = cols[c][6].value
        pwr_con = cols[c][7].value
        dim_val = cols[c][11].value
        ut_val = cols[c][12].value
        
        if sys_val and mod_val and isinstance(cap_val, (int, float)):
            outdoor_units.append({
                "system": sys_val,
                "series": ser_val,
                "model": mod_val,
                "cap_kw": float(cap_val),
                "nominal_cap": str(nom_val or "-"),
                "power_supply": str(pwr_sup or "-"),
                "power_consumption_kw": str(pwr_con or "-"),
                "dimensions": str(dim_val or "-"),
                "unit_type": str(ut_val or "-")
            })
            
    print(f"✅ 成功自 outdoor_units 讀取 {len(outdoor_units)} 台室外機規格數據！")
    
    # 測試 MXM / MXP 驗證規則庫
    mxm_rules = {
        "4MXM110YVLT": {"max_units": 4, "wall_levels": [22, 28, 36, 41, 50, 60, 71, 80, 90], "duct_levels": [22, 28, 36, 41, 50, 60, 71]},
        "3MXM90YVLT": {"max_units": 3, "wall_levels": [22, 28, 36, 41, 50, 60, 71], "duct_levels": [22, 28, 36, 41, 50, 60]},
        "2MXM75YVLT": {"max_units": 2, "wall_levels": [22, 28, 36, 41, 50], "duct_levels": [22, 28, 36, 41, 50]},
        "2MXM56YVLT": {"max_units": 2, "wall_levels": [22, 28, 36], "duct_levels": [22, 28, 36]},
        "2MXP50ZVLT": {"max_units": 2, "wall_levels": [20, 25, 30], "duct_levels": []},
        "2MXP85ZVLT": {"max_units": 2, "wall_levels": [20, 25, 30, 40, 50, 60, 71], "duct_levels": []}
    }
    
    # 建立模擬測試數據 (包含 2 個室外機群組與 1 個一對一單機)
    simulated_payload = {
        "outdoor_groups": [
            {
                "group_id": "group-vrv-1",
                "group_name": "VRV 室外機群組 #1",
                "system_type": "VRV",
                "outdoor_model": "RXYQ10AYLT",
                "outdoor_cap_kw": 28.0,
                "diversity_factor": 90, # 90% 同開率
                "spaces": [
                    {
                        "space_name": "董事長室",
                        "area_m2": 35.48,
                        "area_ping": 10.73,
                        "calc_basis": 550,
                        "total_cooling_demand": 5902,
                        "best_match_model": "FXDQ63NDNVT",
                        "unit_count": 1,
                        "cap_kw": 7.1,
                        "cap_kcal": 6106,
                        "nominal_cap": "63"
                    },
                    {
                        "space_name": "總經理室",
                        "area_m2": 23.20,
                        "area_ping": 7.02,
                        "calc_basis": 550,
                        "total_cooling_demand": 3861,
                        "best_match_model": "FXDQ40NDNVT",
                        "unit_count": 1,
                        "cap_kw": 4.5,
                        "cap_kcal": 3870,
                        "nominal_cap": "40"
                    },
                    {
                        "space_name": "大會議室",
                        "area_m2": 45.00,
                        "area_ping": 13.61,
                        "calc_basis": 600,
                        "total_cooling_demand": 8166,
                        "best_match_model": "FXMQ100PAVT",
                        "unit_count": 1,
                        "cap_kw": 11.2,
                        "cap_kcal": 9632,
                        "nominal_cap": "100"
                    }
                ]
            },
            {
                "group_id": "group-mxm-2",
                "group_name": "家用多聯 Y 系列群組 #2",
                "system_type": "RA",
                "outdoor_model": "4MXM110YVLT",
                "outdoor_cap_kw": 10.5,
                "diversity_factor": 100,
                "spaces": [
                    {
                        "space_name": "主臥室",
                        "area_m2": 20.00,
                        "area_ping": 6.05,
                        "calc_basis": 500,
                        "total_cooling_demand": 3025,
                        "best_match_model": "FTXM36YVLT",
                        "unit_count": 1,
                        "cap_kw": 3.6,
                        "cap_kcal": 3096,
                        "nominal_cap": "36"
                    },
                    {
                        "space_name": "次臥室",
                        "area_m2": 15.00,
                        "area_ping": 4.54,
                        "calc_basis": 500,
                        "total_cooling_demand": 2269,
                        "best_match_model": "FTXM28YVLT",
                        "unit_count": 1,
                        "cap_kw": 2.8,
                        "cap_kcal": 2408,
                        "nominal_cap": "28"
                    }
                ]
            }
        ],
        "ungrouped_spaces": [
            {
                "space_name": "客廳 (一對一獨立機)",
                "area_m2": 33.00,
                "area_ping": 9.98,
                "calc_basis": 650,
                "total_cooling_demand": 6489,
                "best_match_model": "FTXV71UVLT",
                "unit_count": 1,
                "cap_kw": 7.2,
                "cap_kcal": 6192,
                "nominal_cap": "71",
                "outdoor_model": "RXV71UVLT"
            }
        ]
    }
    
    # 測試縱向跨列合併 (openpyxl Rowspan Engine) 繪製
    wb_report = openpyxl.load_workbook(template_path)
    ws_rep = wb_report.active
    
    start_row = 9
    curr_row = start_row
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    
    # 繪製群組空間
    for group in simulated_payload["outdoor_groups"]:
        spaces = group["spaces"]
        group_start_row = curr_row
        
        # 計算群組內總能力與連結率
        sum_indoor_kw = sum(s["cap_kw"] * s["unit_count"] for s in spaces)
        outdoor_kw = group["outdoor_cap_kw"]
        diversity = group["diversity_factor"]
        
        # VRV 連結率 % ＝ (室內機標稱容量總和 / 室外機容量) * 100%
        connection_ratio = round((sum_indoor_kw / outdoor_kw) * 100, 1) if outdoor_kw > 0 else 0.0
        
        for s in spaces:
            ws_rep.cell(row=curr_row, column=4).value = s["space_name"] # D: 室名
            ws_rep.cell(row=curr_row, column=5).value = s["area_m2"]    # E: 面積
            ws_rep.cell(row=curr_row, column=6).value = s["area_ping"]  # F: 坪數
            ws_rep.cell(row=curr_row, column=8).value = s["calc_basis"] # H: 基準
            ws_rep.cell(row=curr_row, column=12).value = round(s["total_cooling_demand"] / 860.0, 2) # L: 總負荷 kW
            ws_rep.cell(row=curr_row, column=13).value = s["total_cooling_demand"] # M: 總負荷 kcal
            
            # 室內機資訊
            ws_rep.cell(row=curr_row, column=15).value = s["best_match_model"] # O: 室內機型號
            ws_rep.cell(row=curr_row, column=16).value = s["unit_count"]       # P: 台數
            ws_rep.cell(row=curr_row, column=17).value = s["cap_kcal"]         # Q: 冷房能力 kcal
            ws_rep.cell(row=curr_row, column=18).value = s["cap_kw"]           # R: 冷房能力 kW
            
            ws_rep.cell(row=curr_row, column=23).value = s["cap_kcal"] * s["unit_count"] # W: 總能力 kcal
            ws_rep.cell(row=curr_row, column=24).value = s["cap_kw"] * s["unit_count"]   # X: 總能力 kW
            
            curr_row += 1
            
        group_end_row = curr_row - 1
        
        # 🎯 縱向跨列合併 (Rowspan Engine)：室外機型號、連結率%、同開率% 橫跨所屬的所有空間列
        if group_end_row > group_start_row:
            # 假設室外機資訊寫在 AE (31), AF (32), AG (33)
            ws_rep.cell(row=group_start_row, column=31).value = group["outdoor_model"]
            ws_rep.cell(row=group_start_row, column=32).value = f"{connection_ratio}%"
            ws_rep.cell(row=group_start_row, column=33).value = f"{diversity}%"
            
            ws_rep.merge_cells(start_row=group_start_row, end_row=group_end_row, start_column=31, end_column=31)
            ws_rep.merge_cells(start_row=group_start_row, end_row=group_end_row, start_column=32, end_column=32)
            ws_rep.merge_cells(start_row=group_start_row, end_row=group_end_row, start_column=33, end_column=33)
            
            # 置中與對齊
            ws_rep.cell(row=group_start_row, column=31).alignment = align_center
            ws_rep.cell(row=group_start_row, column=32).alignment = align_center
            ws_rep.cell(row=group_start_row, column=33).alignment = align_center
            
            print(f"✨ 成功對 {group['group_name']} 執行縱向跨列合併 (Rows {group_start_row} ~ {group_end_row})！連結率: {connection_ratio}%")
        else:
            ws_rep.cell(row=group_start_row, column=31).value = group["outdoor_model"]
            ws_rep.cell(row=group_start_row, column=32).value = f"{connection_ratio}%"
            ws_rep.cell(row=group_start_row, column=33).value = f"{diversity}%"

    # 繪製一對一單機空間
    for s in simulated_payload["ungrouped_spaces"]:
        ws_rep.cell(row=curr_row, column=4).value = s["space_name"]
        ws_rep.cell(row=curr_row, column=5).value = s["area_m2"]
        ws_rep.cell(row=curr_row, column=6).value = s["area_ping"]
        ws_rep.cell(row=curr_row, column=8).value = s["calc_basis"]
        ws_rep.cell(row=curr_row, column=12).value = round(s["total_cooling_demand"] / 860.0, 2)
        ws_rep.cell(row=curr_row, column=13).value = s["total_cooling_demand"]
        
        ws_rep.cell(row=curr_row, column=15).value = s["best_match_model"]
        ws_rep.cell(row=curr_row, column=16).value = s["unit_count"]
        ws_rep.cell(row=curr_row, column=17).value = s["cap_kcal"]
        ws_rep.cell(row=curr_row, column=18).value = s["cap_kw"]
        ws_rep.cell(row=curr_row, column=23).value = s["cap_kcal"] * s["unit_count"]
        ws_rep.cell(row=curr_row, column=24).value = s["cap_kw"] * s["unit_count"]
        
        ws_rep.cell(row=curr_row, column=31).value = s.get("outdoor_model", "一對一室外機")
        ws_rep.cell(row=curr_row, column=32).value = "100%"
        ws_rep.cell(row=curr_row, column=33).value = "100%"
        curr_row += 1
        
    output_test_path = "backend/temp_merged_selection_report.xlsx"
    wb_report.save(output_test_path)
    print(f"🎉 測試報表已成功輸出至: {output_test_path}")

if __name__ == "__main__":
    test_outdoor_db_and_export()
