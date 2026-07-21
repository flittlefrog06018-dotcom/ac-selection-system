import os
import re
import logging
from typing import List, Dict, Any, Tuple
import openpyxl

from app.config import settings

logger = logging.getLogger(__name__)

class CalculationService:
    @classmethod
    def get_db_path(cls) -> str:
        return os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
            "product_database",
            settings.LOAD_DB_NAME
        )

    @classmethod
    def check_and_load_vv16_criteria(cls, sheet_name: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        Loads the criteria guidelines from the spreadsheet '空調負荷基準表.xlsx' (Step 4 of legacy script).
        """
        db_path = cls.get_db_path()
        parsed_rules = []
        modifiers_headers = []

        if not os.path.exists(db_path):
            logger.warning(f"Criteria workbook {settings.LOAD_DB_NAME} not found at {db_path}. Using fallback defaults.")
            return cls._get_fallback_criteria()

        try:
            wb = openpyxl.load_workbook(db_path, data_only=True)
            target_sheet = None
            clean_sheet_name = re.sub(r'[^\w]', '', sheet_name)
            
            # Match sheet name
            for name in wb.sheetnames:
                clean_name = re.sub(r'[^\w]', '', name)
                if clean_name in clean_sheet_name or clean_sheet_name in clean_name:
                    target_sheet = name
                    break
                    
            if not target_sheet:
                for name in wb.sheetnames:
                    if "住宅" in name and "住宅" in sheet_name:
                        target_sheet = name
                        break
                if not target_sheet:
                    target_sheet = wb.sheetnames[0]
                    
            ws = wb[target_sheet]
            
            # Read modifiers from headers (Row 1, Column 3 onwards until '特殊熱源')
            col = 3
            while col <= ws.max_column:
                header_val = ws.cell(row=1, column=col).value
                if header_val is None:
                    col += 1
                    continue
                header_str = str(header_val).strip()
                if "特殊熱源" in header_str:
                    break
                modifiers_headers.append({"col": col, "name": header_str})
                col += 1
                
            # Read rows
            for row in range(2, ws.max_row + 1):
                kw_val = ws.cell(row=row, column=1).value
                base_kcal = ws.cell(row=row, column=2).value
                if not kw_val or base_kcal is None:
                    continue
                
                norm_str = str(kw_val).replace("，", ",").replace("、", ",").replace("\n", ",")
                keywords = [k.strip() for k in norm_str.split(",") if k.strip()]
                
                row_modifiers = {}
                for h in modifiers_headers:
                    cell_val = ws.cell(row=row, column=h["col"]).value
                    try:
                        if cell_val is None:
                            val_float = 0.0
                        elif isinstance(cell_val, str) and "%" in cell_val:
                            val_float = float(cell_val.replace("%", "").strip()) / 100.0
                        else:
                            val_float = float(cell_val)
                    except:
                        val_float = 0.0
                    row_modifiers[h["name"]] = val_float
                    
                parsed_rules.append({
                    "keywords": keywords,
                    "base_kcal": float(base_kcal),
                    "modifiers": row_modifiers
                })
                
            return parsed_rules, modifiers_headers
            
        except Exception as e:
            logger.error(f"Error loading criteria sheet {sheet_name}: {e}. Returning fallback.")
            return cls._get_fallback_criteria()

    @staticmethod
    def _get_fallback_criteria() -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        Default backup criteria if criteria sheet doesn't exist.
        """
        default_headers = [
            {"col": 3, "name": "全內周"},
            {"col": 4, "name": "2面牆"},
            {"col": 5, "name": "西曬"},
            {"col": 6, "name": "挑高"},
            {"col": 7, "name": "頂曬"}
        ]
        
        parsed_rules = [
            {
                "keywords": ["臥室", "主臥室", "客房", "更衣間", "更衣室", "浴室", "主臥浴室", "廁所"],
                "base_kcal": 500.0,
                "modifiers": {"全內周": -0.10, "2面牆": 0.05, "西曬": 0.06, "挑高": 0.04, "頂曬": 0.05}
            },
            {
                "keywords": ["客廳", "餐廳", "飯廳", "吧台區", "合約洽談區", "視訊室兼餐廳", "會客大廳"],
                "base_kcal": 550.0,
                "modifiers": {"全內周": -0.10, "2面牆": 0.05, "西曬": 0.06, "挑高": 0.04, "頂曬": 0.05}
            },
            {
                "keywords": ["辦公室", "董事長室", "總經理室", "開放辦公區", "會議室", "A會議室"],
                "base_kcal": 600.0,
                "modifiers": {"全內周": -0.10, "2面牆": 0.05, "西曬": 0.06, "挑高": 0.04, "頂曬": 0.05}
            },
            {
                "keywords": ["機房", "系統主機房"],
                "base_kcal": 700.0,
                "modifiers": {"全內周": -0.10, "2面牆": 0.05, "西曬": 0.06, "挑高": 0.04, "頂曬": 0.05}
            }
        ]
        return parsed_rules, default_headers

    @staticmethod
    def match_space_vv16_data(space_name: str, 
                              parsed_rules: List[Dict[str, Any]], 
                              default_kcal: float = 500.0, 
                              modifiers_headers: List[Dict[str, Any]] = None) -> Tuple[float, Dict[str, float], bool]:
        """
        Fuzzy matches a room name against rules to fetch its base kcal/h and modifier ratios (from script).
        """
        cleaned_name = space_name.strip()
        base_search_name = re.sub(r'[\d一二三四五六七八九十]+$', '', cleaned_name)
        if not base_search_name:
            base_search_name = cleaned_name

        # 1. Exact or base name match
        for rule in parsed_rules:
            for kw in rule["keywords"]:
                if kw == cleaned_name or kw == base_search_name:
                    return rule["base_kcal"], rule["modifiers"], True
                    
        # 2. Substring fuzzy match
        for rule in parsed_rules:
            for kw in rule["keywords"]:
                if kw in cleaned_name or cleaned_name in kw or kw in base_search_name or base_search_name in kw:
                    return rule["base_kcal"], rule["modifiers"], True
                    
        # 3. Default fallback
        if parsed_rules:
            fallback_mods = parsed_rules[0]["modifiers"].copy()
        else:
            fallback_mods = {h["name"]: 0.0 for h in (modifiers_headers or [])}
            
        return default_kcal, fallback_mods, False

    @classmethod
    def process_and_calculate_loads(cls, 
                                    rooms: List[Dict[str, Any]], 
                                    project_sheet_name: str) -> Tuple[List[Dict[str, Any]], List[str]]:
        """
        Core Calculation Engine. Calculates:
        - area (m2 and ping)
        - base thermal load matching
        - modifier additions
        - special kW load conversion
        - total kW/kcal loads
        """
        parsed_rules, modifiers_headers = cls.check_and_load_vv16_criteria(project_sheet_name)
        modifier_names = [h["name"] for h in modifiers_headers]
        
        processed_rooms = []
        for room in rooms:
            name = room.get("room_name", "未知空間").strip()
            
            # Read area and unit
            raw_val = float(room.get("area_raw") if room.get("area_raw") is not None else room.get("area_m2", 10.0))
            unit = room.get("unit", "m2").strip()
            
            # Calculate area values (conversion rule from script)
            if unit in ["P", "坪"]:
                ping_val = round(raw_val, 1)
                area_m2 = round(ping_val / 0.3025, 1)
            else:
                area_m2 = round(raw_val, 1)
                ping_val = round(area_m2 * 0.3025, 1)
                
            # Match rules
            base_kcal, r_mods, matched = cls.match_space_vv16_data(
                name, parsed_rules, default_kcal=500.0, modifiers_headers=modifiers_headers
            )
            
            # Sum up checkbox percentages
            percentage_sum = 0.0
            checks = room.get("checks", {})
            for h_name in modifier_names:
                # If checked (True) in frontend
                if checks.get(h_name, False):
                    percentage_sum += r_mods.get(h_name, 0.0)
                    
            # Parse special_kw
            special_kw = float(room.get("special_kw", 0.0) or 0.0)
            special_total_kcal = special_kw * 860.0
            
            # Formula (from your script):
            # adjusted_base_kcal = base_kcal * (1 + percentage_sum)
            # special_kcal_per_ping = special_total_kcal / ping_val
            # final_suggested_kcal_per_ping = adjusted_base_kcal + special_kcal_per_ping
            adjusted_base_kcal = base_kcal * (1 + percentage_sum)
            special_kcal_per_ping = (special_total_kcal / ping_val) if ping_val > 0 else 0.0
            
            final_suggested_kcal_per_ping = round(adjusted_base_kcal + special_kcal_per_ping, 1)
            
            kw_per_ping = round(final_suggested_kcal_per_ping / 860.0, 2)
            total_load_kw = round(ping_val * kw_per_ping, 1)
            total_load_kcal = round(ping_val * final_suggested_kcal_per_ping, 1)
            total_load_w = round(total_load_kw * 1000.0, 1)
            
            # Store values in dictionary to return
            processed_rooms.append({
                "room_name": name,
                "area_raw": raw_val,
                "area_m2": area_m2,
                "ping_val": ping_val,
                "unit": unit,
                "base_kcal": base_kcal,
                "matched": matched,
                # Store dynamic modifiers values for current room to let frontend know individual rates
                "modifier_rates": r_mods,
                "checks": checks,
                "special_kw": special_kw,
                
                # Calculation results
                "final_suggested_kcal_per_ping": final_suggested_kcal_per_ping,
                "kw_per_ping": kw_per_ping,
                "total_load_kw": total_load_kw,
                "total_load_kcal": total_load_kcal,
                "total_load_w": total_load_w,
                
                # Geometrics (for sorting)
                "center_x": float(room.get("center_x", 0.5)),
                "center_y": float(room.get("center_y", 0.5)),
            })
            
        return processed_rooms, modifier_names
