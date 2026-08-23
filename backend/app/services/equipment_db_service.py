import os
import logging
from typing import List, Dict, Any, Optional, Tuple, Set
import openpyxl

logger = logging.getLogger(__name__)

class EquipmentDBService:
    _instance = None
    
    def __init__(self):
        self.units: List[Dict[str, Any]] = []
        self.is_loaded = False
        self.db_path: Optional[str] = None
        
    @classmethod
    def get_instance(cls):
        if cls._instance is None:
            cls._instance = EquipmentDBService()
        return cls._instance

    @staticmethod
    def get_candidate_db_paths() -> List[str]:
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        root_dir = os.path.dirname(base_dir)
        return [
            os.path.join(base_dir, "product_database", "EQUIPMENT_Data.xlsx"),
            os.path.join(base_dir, "product_database", "EQUIPMENT_DB.xlsx"),
            os.path.join(root_dir, "EQUIPMENT_Data.xlsx"),
            os.path.join(root_dir, "EQUIPMENT_DB.xlsx"),
            os.path.join(base_dir, "EQUIPMENT_Data.xlsx"),
            os.path.join(base_dir, "EQUIPMENT_DB.xlsx"),
            os.path.abspath("backend/product_database/EQUIPMENT_Data.xlsx"),
            os.path.abspath("EQUIPMENT_Data.xlsx"),
            os.path.abspath("EQUIPMENT_DB.xlsx"),
        ]

    def load_equipment_db(self, raise_error_on_missing: bool = False) -> bool:
        """
        Reads 'indoor_units' sheet from EQUIPMENT_Data.xlsx (E2:E12 rightwards).
        Row 2: System (RA, SA, VRV)
        Row 3: Series (橫綱X系列, etc.)
        Row 4: Model (FTXV22XVLT, etc.)
        Row 5: Cap kW (2.2, etc.)
        Row 6: Nominal cap / capacity index (-)
        Row 7: Power supply (室外機供電, etc.)
        Row 8: Power consumption kW (-)
        Row 11: Dimensions mm HxWxD (299x838x275, etc.)
        Row 12: Indoor unit type (壁掛式, 吊隱式, 嵌入式, 天吊式, 箱型機, etc.)
        """
        candidate_paths = self.get_candidate_db_paths()
        target_path = None
        for p in candidate_paths:
            if os.path.exists(p):
                target_path = p
                break
                
        if not target_path:
            err_msg = f"❌ 找不到設備資料庫檔案 EQUIPMENT_Data.xlsx！已搜尋路徑: {candidate_paths[:3]}"
            logger.error(err_msg)
            if raise_error_on_missing:
                raise FileNotFoundError(err_msg)
            return False

        try:
            wb = openpyxl.load_workbook(target_path, data_only=True)
            sheet_names = wb.sheetnames
            
            target_sheet_name = None
            for name in sheet_names:
                if "indoor" in name.lower() or "室內機" in name:
                    target_sheet_name = name
                    break
            if not target_sheet_name:
                target_sheet_name = sheet_names[0]
                
            ws = wb[target_sheet_name]
            units_list = []
            
            # Start from Column E (col 5) rightwards
            max_col = ws.max_column
            for col in range(5, max_col + 1):
                raw_system = ws.cell(row=2, column=col).value
                raw_series = ws.cell(row=3, column=col).value
                raw_model = ws.cell(row=4, column=col).value
                raw_cap_kw = ws.cell(row=5, column=col).value
                raw_nominal = ws.cell(row=6, column=col).value
                raw_power_sup = ws.cell(row=7, column=col).value
                raw_power_con = ws.cell(row=8, column=col).value
                raw_current = ws.cell(row=9, column=col).value
                raw_dim = ws.cell(row=11, column=col).value
                raw_type = ws.cell(row=12, column=col).value
                
                # Check if model or system is empty (stop condition)
                if not raw_model and not raw_cap_kw and not raw_system:
                    continue
                    
                model_str = str(raw_model).strip().upper() if raw_model else ""
                if not model_str or model_str == "NONE":
                    continue
                    
                system_str = str(raw_system).strip().upper() if raw_system else "VRV"
                series_str = str(raw_series).strip() if raw_series else "標準系列"
                unit_type_str = str(raw_type).strip() if raw_type else "壁掛式"
                
                try:
                    cap_kw_val = float(raw_cap_kw)
                except (ValueError, TypeError):
                    cap_kw_val = 2.2
                    
                def clean_raw_val(v):
                    if v is None:
                        return "-"
                    s = str(v).strip()
                    return s if s else "-"

                unit_obj = {
                    "system": system_str,
                    "series": series_str,
                    "model": model_str,
                    "cap_kw": cap_kw_val,
                    "nominal_cap": clean_raw_val(raw_nominal),
                    "power_supply": clean_raw_val(raw_power_sup),
                    "power_consumption_kw": clean_raw_val(raw_power_con),
                    "mca": clean_raw_val(raw_current),
                    "dimensions": clean_raw_val(raw_dim),
                    "unit_type": unit_type_str,
                    "col_index": col
                }
                units_list.append(unit_obj)
                
            self.units = units_list
            self._indoor_units_map = {u["model"].upper(): u for u in units_list}
            self.load_outdoor_units(target_path)
            self.is_loaded = True
            self.db_path = target_path
            logger.info(f"✨ 成功從 {target_path} 讀取 {len(units_list)} 台室內機與 {len(self.outdoor_units)} 台室外機規格數據！")
            return True
        except Exception as e:
            logger.error(f"Failed to parse equipment DB {target_path}: {e}")
            if raise_error_on_missing:
                raise e
            return False

    def load_outdoor_units(self, file_path: str):
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            if "outdoor_units" not in wb.sheetnames:
                self.outdoor_units = []
                self._outdoor_units_map = {}
                return
            ws = wb["outdoor_units"]
            cols = list(ws.columns)
            outdoors = []
            outdoors_map = {}
            for c in range(2, len(cols)):
                sys_val = cols[c][1].value   # row 2
                ser_val = cols[c][2].value   # row 3
                mod_val = cols[c][3].value   # row 4
                cap_val = cols[c][4].value   # row 5
                nom_val = cols[c][5].value   # row 6
                pwr_sup = cols[c][6].value   # row 7
                pwr_con = cols[c][7].value   # row 8
                mca_val = cols[c][8].value   # row 9 (電路最大電流 MCA)
                mfa_val = cols[c][9].value   # row 10 (保險絲最大電流 MFA)
                dim_val = cols[c][11].value  # row 12 (尺寸 mm HxWxD)
                ut_val = cols[c][12].value   # row 13 (型式)
                
                if sys_val and mod_val and isinstance(cap_val, (int, float)):
                    def clean_val(v):
                        if v is None:
                            return "-"
                        s = str(v).strip()
                        return s if s else "-"

                    m_str = str(mod_val).strip().upper()
                    obj = {
                        "system": str(sys_val).strip(),
                        "series": str(ser_val or "").strip(),
                        "model": m_str,
                        "cap_kw": float(cap_val),
                        "nominal_cap": clean_val(nom_val),
                        "power_supply": clean_val(pwr_sup),
                        "power_consumption_kw": clean_val(pwr_con),
                        "mca": clean_val(mca_val),
                        "mfa": clean_val(mfa_val),
                        "dimensions": clean_val(dim_val),
                        "unit_type": clean_val(ut_val)
                    }
                    outdoors.append(obj)
                    outdoors_map[m_str] = obj
            self.outdoor_units = outdoors
            self._outdoor_units_map = outdoors_map
        except Exception as e:
            logger.error(f"Failed to parse outdoor_units sheet: {e}")
            self.outdoor_units = []
            self._outdoor_units_map = {}

    def get_indoor_unit_info(self, model_name: str) -> Optional[Dict[str, Any]]:
        if not self.is_loaded:
            self.load_equipment_db()
        if not model_name:
            return None
        m_upper = str(model_name).strip().upper()
        if not hasattr(self, "_indoor_units_map") or not self._indoor_units_map:
            self._indoor_units_map = {u["model"].upper(): u for u in self.units}
        return self._indoor_units_map.get(m_upper)

    def get_outdoor_unit_info(self, model_name: str) -> Optional[Dict[str, Any]]:
        if not self.is_loaded:
            self.load_equipment_db()
        if not model_name:
            return None
        m_upper = str(model_name).strip().upper()
        if hasattr(self, "_outdoor_units_map") and m_upper in self._outdoor_units_map:
            return self._outdoor_units_map[m_upper]
        for u in getattr(self, "outdoor_units", []):
            if u["model"].upper() == m_upper:
                return u
        return None

    def get_systems(self) -> List[str]:
        if not self.is_loaded:
            self.load_equipment_db()
        systems = list(dict.fromkeys([u["system"] for u in self.units if u.get("system")]))
        return systems if systems else ["RA", "SA", "VRV"]

    def get_unit_types(self, system: Optional[str] = None) -> List[str]:
        if not self.is_loaded:
            self.load_equipment_db()
        pool = self.units
        if system:
            sys_upper = system.strip().upper()
            pool = [u for u in pool if u["system"] == sys_upper]
        types = list(dict.fromkeys([u["unit_type"] for u in pool if u.get("unit_type")]))
        return types if types else ["壁掛式", "吊隱式", "嵌入式", "天吊式", "箱型機"]

    def get_series(self, system: Optional[str] = None, unit_type: Optional[str] = None) -> List[str]:
        if not self.is_loaded:
            self.load_equipment_db()
        pool = self.units
        if system:
            sys_upper = system.strip().upper()
            pool = [u for u in pool if u["system"] == sys_upper]
        if unit_type:
            ut_str = unit_type.strip()
            pool = [u for u in pool if u["unit_type"] == ut_str]
        series_list = list(dict.fromkeys([u["series"] for u in pool if u.get("series")]))
        return series_list if series_list else ["標準系列"]

    def get_models_filtered(self, system: Optional[str] = None, unit_type: Optional[str] = None, series: Optional[str] = None, target_kw: Optional[float] = None) -> List[Dict[str, Any]]:
        """
        Returns models matching system/type/series, and filters within target_kw +/- 20% if target_kw is provided.
        """
        if not self.is_loaded:
            self.load_equipment_db()
        pool = self.units
        if system:
            sys_upper = system.strip().upper()
            pool = [u for u in pool if u["system"] == sys_upper]
        if unit_type:
            ut_str = unit_type.strip()
            pool = [u for u in pool if u["unit_type"] == ut_str]
        if series:
            ser_str = series.strip()
            pool = [u for u in pool if u["series"] == ser_str]
            
        if not pool:
            return []
            
        if target_kw and target_kw > 0:
            min_kw = target_kw * 0.8
            max_kw = target_kw * 1.2
            filtered = [u for u in pool if min_kw <= u["cap_kw"] <= max_kw]
            if filtered:
                return sorted(filtered, key=lambda x: x["cap_kw"])
                
        return sorted(pool, key=lambda x: x["cap_kw"])

    def auto_select_indoor_unit_dynamic(self, total_load_kw: float, system: str = "VRV", unit_type: Optional[str] = None, series: Optional[str] = None) -> Dict[str, Any]:
        """
        Dynamic indoor unit selection with capacity overflow expansion (Qty > 1) and +/- 20% candidate filtering.
        """
        if not self.is_loaded:
            self.load_equipment_db()

        sys_upper = system.strip().upper()
        # Fallback mapping for system names
        if "住宅" in sys_upper or "RA" in sys_upper or "家用" in sys_upper:
            sys_upper = "RA"
        elif "商用" in sys_upper or "SA" in sys_upper or "商業" in sys_upper:
            sys_upper = "SA"
        elif "VRV" in sys_upper:
            sys_upper = "VRV"
            
        candidates = [u for u in self.units if u["system"] == sys_upper]
        if series and series.strip():
            ser_filtered = [u for u in candidates if u["series"] == series.strip()]
            if ser_filtered:
                candidates = ser_filtered

        if unit_type and unit_type.strip():
            ut_filtered = [u for u in candidates if u["unit_type"] == unit_type.strip()]
            if ut_filtered:
                candidates = ut_filtered
                
        if not candidates:
            # Fallback to any VRV or any unit
            candidates = self.units if self.units else [
                {"model": "FXSQ28PAVT", "cap_kw": 2.8, "system": "VRV", "unit_type": "壁掛式", "series": "標準", "nominal_cap": "-", "power_supply": "-", "power_consumption_kw": "-", "dimensions": "-"}
            ]

        candidates_sorted = sorted(candidates, key=lambda x: x["cap_kw"])
        max_unit_cap = candidates_sorted[-1]["cap_kw"]
        
        # Calculate required unit count (Qty) if demand exceeds max single unit
        if total_load_kw > max_unit_cap:
            qty = int(round((total_load_kw / max_unit_cap) + 0.499, 0))
            qty = max(1, qty)
        else:
            qty = 1
            
        target_single_kw = total_load_kw / qty
        
        # +/- 20% candidate range filter
        min_kw = target_single_kw * 0.8
        max_kw = target_single_kw * 1.2
        range_candidates = [u for u in candidates_sorted if min_kw <= u["cap_kw"] <= max_kw]
        
        selected_unit = None
        if range_candidates:
            # Pick the smallest capacity in +/- 20% range that meets target_single_kw
            meets = [u for u in range_candidates if u["cap_kw"] >= target_single_kw]
            selected_unit = meets[0] if meets else range_candidates[-1]
        else:
            # Fallback: pick smallest capacity >= target_single_kw in entire candidate list
            meets = [u for u in candidates_sorted if u["cap_kw"] >= target_single_kw]
            selected_unit = meets[0] if meets else candidates_sorted[-1]
            
        cap_kw = selected_unit["cap_kw"]
        cap_kcal = round(cap_kw * 860.0, 1)
        
        return {
            "model": selected_unit["model"],
            "qty": qty,
            "cap_kw": cap_kw,
            "cap_kcal": cap_kcal,
            "total_cap_kw": round(qty * cap_kw, 1),
            "total_cap_kcal": round(qty * cap_kcal, 1),
            "nominal_cap": selected_unit.get("nominal_cap", "-"),
            "power_supply": selected_unit.get("power_supply", "-"),
            "power_consumption_kw": selected_unit.get("power_consumption_kw", "-"),
            "dimensions": selected_unit.get("dimensions", "-"),
            "unit_type": selected_unit.get("unit_type", "壁掛式"),
            "series": selected_unit.get("series", "標準系列"),
            "system": selected_unit.get("system", "VRV")
        }
