import os
import io
import logging
from typing import List, Dict, Any
import openpyxl

from app.config import settings

logger = logging.getLogger(__name__)

class ExportService:
    @classmethod
    def get_template_path(cls) -> str:
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        root_dir = os.path.dirname(base_dir)
        candidates = [
            os.path.join(base_dir, "product_database", settings.TEMPLATE_NAME),
            os.path.join(base_dir, settings.TEMPLATE_NAME),
            os.path.join(root_dir, settings.TEMPLATE_NAME),
            os.path.join(root_dir, "backend", "product_database", settings.TEMPLATE_NAME),
            os.path.abspath("backend/product_database/選機表-.xlsx"),
            os.path.abspath("product_database/選機表-.xlsx"),
            os.path.abspath("選機表-.xlsx")
        ]
        for c in candidates:
            if os.path.exists(c):
                return c
        return candidates[0]

    @classmethod
    def generate_excel_report(cls, rooms_data: List[Dict[str, Any]], outdoor_groups: List[Dict[str, Any]] = None) -> io.BytesIO:
        """
        Loads the template Excel sheet '選機表-.xlsx', maps properties,
        and performs vertical cell merging (rowspan) for outdoor unit cards.
        """
        template_path = cls.get_template_path()
        
        if not os.path.exists(template_path):
            logger.warning(f"Excel template file {settings.TEMPLATE_NAME} not found at {template_path}. Generating simple output.")
            return cls._generate_fallback_excel(rooms_data)
            
        try:
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            from openpyxl.styles import Alignment
            align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            start_row = settings.START_ROW
            
            # Flatten rooms based on outdoor_groups if provided, else use rooms_data directly
            flat_rows_to_render = []
            group_spans = [] # List of outdoor group render descriptors
            
            from app.services.equipment_db_service import EquipmentDBService
            db_service = EquipmentDBService.get_instance()

            if outdoor_groups and len(outdoor_groups) > 0:
                current_idx = 0
                for group in outdoor_groups:
                    group_spaces = group.get("spaces", [])
                    if not group_spaces:
                        continue
                    g_start = start_row + current_idx
                    g_end = g_start + len(group_spaces) - 1
                    
                    sum_indoor_kw = sum(float(s.get("indoor_capacity_kw", s.get("cap_kw", 0.0))) * int(s.get("qty", s.get("unit_count", 1))) for s in group_spaces)
                    outdoor_model = str(group.get("outdoor_model", "室外機")).strip()
                    outdoor_qty = int(group.get("outdoor_qty", 1))
                    
                    # Look up outdoor unit specs from EQUIPMENT_Data (outdoor_units sheet)
                    outdoor_info = db_service.get_outdoor_unit_info(outdoor_model)
                    
                    outdoor_kw = float(group.get("outdoor_cap_kw", 0.0))
                    if outdoor_kw <= 0 and outdoor_info:
                        outdoor_kw = float(outdoor_info.get("cap_kw", 0.0))
                        
                    conn_ratio = round((sum_indoor_kw / outdoor_kw) * 100, 1) if outdoor_kw > 0 else 0.0
                    conn_ratio_str = f"{conn_ratio}%" if outdoor_kw > 0 else "-"
                    
                    group_spans.append({
                        "start_row": g_start,
                        "end_row": g_end,
                        "outdoor_model": outdoor_model,
                        "outdoor_qty": outdoor_qty,
                        "conn_ratio_str": conn_ratio_str,
                        "outdoor_info": outdoor_info,
                        "fallback_cap_kw": outdoor_kw
                    })
                    
                    for s in group_spaces:
                        s["_in_group"] = True
                        flat_rows_to_render.append(s)
                        current_idx += 1
                        
                # Add any remaining ungrouped spaces
                if rooms_data:
                    rendered_names = {s.get("room_name", s.get("space_name", "")) for s in flat_rows_to_render}
                    for r in rooms_data:
                        r_name = r.get("room_name", r.get("space_name", ""))
                        if r_name not in rendered_names:
                            r["_in_group"] = False
                            flat_rows_to_render.append(r)
                            current_idx += 1
            else:
                flat_rows_to_render = rooms_data or []
                for r in flat_rows_to_render:
                    r["_in_group"] = False

            # If the number of spaces exceeds the template placeholder rows, insert rows dynamically
            template_rows = settings.TEMPLATE_ROWS
            if len(flat_rows_to_render) > template_rows:
                for _ in range(len(flat_rows_to_render) - template_rows):
                    ws.insert_rows(start_row + template_rows - 1)
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=start_row + template_rows - 1, column=c)._style = ws.cell(row=start_row, column=c)._style
                        
            # Write rooms data
            for i, room in enumerate(flat_rows_to_render):
                row_idx = start_row + i
                
                name = room.get("room_name", room.get("space_name", "")).strip()
                area_m2 = float(room.get("area_m2", 0.0))
                ping_val = float(room.get("ping_val", room.get("area_ping", 0.0)))
                
                final_suggested_kcal_per_ping = float(room.get("final_suggested_kcal_per_ping", room.get("calc_basis", 500.0)))
                kw_per_ping = float(room.get("kw_per_ping", round(final_suggested_kcal_per_ping / 860.0, 2)))
                total_load_kcal = float(room.get("total_load_kcal", room.get("total_cooling_demand", round(ping_val * final_suggested_kcal_per_ping))))
                total_load_kw = float(room.get("total_load_kw", round(total_load_kcal / 860.0, 2)))
                
                matched_model = room.get("indoor_model", room.get("best_match_model", "")).strip()
                system_type = str(room.get("system_type", room.get("system", "VRV"))).strip().upper()
                qty = int(room.get("qty", room.get("unit_count", 1)))
                cap_kw = float(room.get("indoor_capacity_kw", room.get("cap_kw", 0.0)))
                cap_kcal = float(room.get("indoor_capacity_kcal", round(cap_kw * 860.0, 1)))
                
                # 🎯 載入室內機 EQUIPMENT_Data 規格資訊
                indoor_info = db_service.get_indoor_unit_info(matched_model)
                
                # 🎯 經理指定室內機電源判斷規則：
                # 只有室內機為 VRV 系統，或是商用 1 對 1 的風管型 (型號是 FBA) 時會需要使用獨立電源 (1φ, 220V, 60Hz)，
                # 其餘的室內機皆使用外機供電，所以在室內機的電源欄位中以 "-" 標示即可。
                matched_upper = matched_model.upper()
                if "VRV" in system_type or matched_upper.startswith("FX") or matched_upper.startswith("FBA"):
                    power_supply = "1φ, 220V, 60Hz"
                else:
                    power_supply = "-"
                    
                # 🎯 標稱能力：帶入 EQUIPMENT_Data 的 indoor_units 第 6 列
                raw_nom = indoor_info.get("nominal_cap") if indoor_info else None
                if raw_nom is not None and str(raw_nom).strip() not in ["-", "None", ""]:
                    try:
                        if "." in str(raw_nom):
                            nominal_cap_val = float(raw_nom)
                        else:
                            nominal_cap_val = int(raw_nom)
                    except (ValueError, TypeError):
                        nominal_cap_val = str(raw_nom).strip()
                else:
                    nominal_cap_val = cap_kw

                power_consumption_kw = room.get("power_consumption_kw") if room.get("power_consumption_kw") and room.get("power_consumption_kw") != "-" else (indoor_info.get("power_consumption_kw", "-") if indoor_info else "-")
                indoor_current_a = indoor_info.get("mca", "-") if indoor_info else "-"
                dimensions = room.get("dimensions") if room.get("dimensions") and room.get("dimensions") != "-" else (indoor_info.get("dimensions", "-") if indoor_info else "-")

                # Write to exact column mappings (Aligned with user screenshot)
                ws.cell(row=row_idx, column=settings.NAME_COL).value = name             # D: 室名
                ws.cell(row=row_idx, column=settings.AREA_COL).value = area_m2          # E: 面積 ㎡
                ws.cell(row=row_idx, column=settings.PING_COL).value = ping_val         # F: 坪數 P
                ws.cell(row=row_idx, column=settings.LOAD_H_COL).value = final_suggested_kcal_per_ping # H: 建議負荷
                
                cell_k = ws.cell(row=row_idx, column=settings.LOAD_K_COL)
                cell_k.value = kw_per_ping
                cell_k.number_format = '0.00'
                
                ws.cell(row=row_idx, column=settings.LOAD_L_COL).value = total_load_kw   # L: 總負荷 kW
                ws.cell(row=row_idx, column=settings.LOAD_M_COL).value = total_load_kcal # M: 總負荷 kcal
                
                # 🎯 室內機修正欄位 (N ~ V)
                ws.cell(row=row_idx, column=14).value = matched_model            # N: 室內機型號
                ws.cell(row=row_idx, column=15).value = qty                      # O: 室內機台數
                ws.cell(row=row_idx, column=16).value = cap_kcal                 # P: 冷房能力 (kcal/hr)
                ws.cell(row=row_idx, column=17).value = cap_kw                   # Q: 冷房能力 (kW)
                ws.cell(row=row_idx, column=18).value = nominal_cap_val          # R: 標稱能力 (EQUIPMENT_Data 第6列)
                ws.cell(row=row_idx, column=19).value = power_supply             # S: 電源
                ws.cell(row=row_idx, column=20).value = power_consumption_kw     # T: 單台耗電量 (kW)
                ws.cell(row=row_idx, column=21).value = indoor_current_a         # U: 單台最大電流 (A)
                ws.cell(row=row_idx, column=22).value = dimensions               # V: 尺寸 mm (H×W×D)
                
                ws.cell(row=row_idx, column=settings.TOTAL_KCAL_W_COL).value = float(qty * cap_kcal)
                ws.cell(row=row_idx, column=settings.TOTAL_KW_X_COL).value = float(qty * cap_kw)
                
                if ping_val > 0:
                    ws.cell(row=row_idx, column=settings.PER_PING_KCAL_AB_COL).value = int(round(cap_kcal / ping_val, 0))
                    ws.cell(row=row_idx, column=settings.PER_PING_KW_AC_COL).value = round(cap_kw / ping_val, 1)
                
                if (qty * cap_kw) > 0:
                    ws.cell(row=row_idx, column=settings.PING_PER_USRT_AD_COL).value = round(ping_val / ((qty * cap_kw) / 3.516), 1)

                # 🎯 單機/未分組空間之室外機欄位直接填入 (AE ~ AO)
                if not room.get("_in_group", False):
                    out_model = room.get("outdoor_model", "").strip()
                    if not out_model:
                        # 嘗試由內機型號或衍生名推導
                        if matched_upper.startswith("FTXM"):
                            out_model = matched_upper.replace("FTXM", "RXM")
                        elif matched_upper.startswith("FTXV"):
                            out_model = matched_upper.replace("FTXV", "RXV")
                        elif matched_upper.startswith("FBA"):
                            out_model = matched_upper.replace("FBA", "RZA")
                        else:
                            out_model = matched_upper
                            
                    out_info = db_service.get_outdoor_unit_info(out_model)
                    out_cap_kw = float(out_info.get("cap_kw", cap_kw)) if out_info else cap_kw
                    out_cap_kcal = round(out_cap_kw * 860.0, 1) if out_cap_kw > 0 else "-"
                    out_nominal = out_info.get("nominal_cap", "-") if out_info else "-"
                    out_pwr_con = out_info.get("power_consumption_kw", "-") if out_info else "-"
                    out_pwr_sup = out_info.get("power_supply", "-") if out_info else "-"
                    out_mca = out_info.get("mca", "-") if out_info else "-"
                    out_mfa = out_info.get("mfa", "-") if out_info else "-"
                    out_dim = out_info.get("dimensions", "-") if out_info else "-"

                    ws.cell(row=row_idx, column=31).value = out_model if out_model else "-"   # AE: 室外機型號 (第4列)
                    ws.cell(row=row_idx, column=32).value = int(room.get("outdoor_qty", 1))   # AF: 室外機台數 (以選型台數為主)
                    ws.cell(row=row_idx, column=33).value = out_cap_kcal                     # AG: 冷房能力 (kcal/hr)
                    ws.cell(row=row_idx, column=34).value = out_cap_kw                       # AH: 冷房能力 (kW) (第5列)
                    ws.cell(row=row_idx, column=35).value = out_nominal                      # AI: 標稱能力 (第6列)
                    ws.cell(row=row_idx, column=36).value = "100%"                           # AJ: 連結率 % (以選型計算結果為主)
                    ws.cell(row=row_idx, column=37).value = out_pwr_con                      # AK: 耗電量 (kW) (第8列)
                    ws.cell(row=row_idx, column=38).value = out_pwr_sup                      # AL: 電源 (第7列)
                    ws.cell(row=row_idx, column=39).value = out_mca                          # AM: 電路最大電流 (A) (第9列)
                    ws.cell(row=row_idx, column=40).value = out_mfa                          # AN: 保險絲最大電流 (A) (第10列)
                    ws.cell(row=row_idx, column=41).value = out_dim                          # AO: 尺寸 mm (H×W×D) (第12列)

            # 🎯 執行室外機群組縱向跨列合併與 EQUIPMENT_Data 對應填入 (openpyxl Rowspan Engine)
            for span in group_spans:
                s_r = span["start_row"]
                e_r = span["end_row"]
                out_info = span["outdoor_info"]
                
                out_model = span["outdoor_model"]
                out_qty = span["outdoor_qty"]
                conn_ratio_str = span["conn_ratio_str"]
                
                out_cap_kw = float(out_info.get("cap_kw", span["fallback_cap_kw"])) if out_info else span["fallback_cap_kw"]
                out_cap_kcal = round(out_cap_kw * 860.0, 1) if out_cap_kw > 0 else "-"
                out_nominal = out_info.get("nominal_cap", "-") if out_info else "-"
                out_pwr_con = out_info.get("power_consumption_kw", "-") if out_info else "-"
                out_pwr_sup = out_info.get("power_supply", "-") if out_info else "-"
                out_mca = out_info.get("mca", "-") if out_info else "-"
                out_mfa = out_info.get("mfa", "-") if out_info else "-"
                out_dim = out_info.get("dimensions", "-") if out_info else "-"

                # 填寫室外機專屬 11 個欄位 (AE ~ AO, Columns 31 ~ 41)
                ws.cell(row=s_r, column=31).value = out_model       # AE: 室外機型號 (第4列)
                ws.cell(row=s_r, column=32).value = out_qty         # AF: 室外機台數 (以選型台數為主)
                ws.cell(row=s_r, column=33).value = out_cap_kcal     # AG: 冷房能力 (kcal/hr)
                ws.cell(row=s_r, column=34).value = out_cap_kw       # AH: 冷房能力 (kW) (第5列)
                ws.cell(row=s_r, column=35).value = out_nominal      # AI: 標稱能力 (第6列)
                ws.cell(row=s_r, column=36).value = conn_ratio_str   # AJ: 連結率 % (以選型計算結果為主)
                ws.cell(row=s_r, column=37).value = out_pwr_con      # AK: 耗電量 (kW) (第8列)
                ws.cell(row=s_r, column=38).value = out_pwr_sup      # AL: 電源 (第7列)
                ws.cell(row=s_r, column=39).value = out_mca          # AM: 電路最大電流 (A) (第9列)
                ws.cell(row=s_r, column=40).value = out_mfa          # AN: 保險絲最大電流 (A) (第10列)
                ws.cell(row=s_r, column=41).value = out_dim          # AO: 尺寸 mm (H×W×D) (第12列)
                
                # 縱向跨列合併對齊
                for col_c in range(31, 42):
                    if e_r > s_r:
                        ws.merge_cells(start_row=s_r, end_row=e_r, start_column=col_c, end_column=col_c)
                    ws.cell(row=s_r, column=col_c).alignment = align_center

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output
            
        except Exception as e:
            logger.error(f"Failed to export using Excel template: {e}")
            return cls._generate_fallback_excel(rooms_data)

    @classmethod
    def _generate_fallback_excel(cls, rooms_data: List[Dict[str, Any]]) -> io.BytesIO:
        """
        Creates a basic spreadsheet if the template is not present.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "選機結果表"
        
        headers = ["空間名稱", "面積 (m2)", "坪數", "每坪kcal", "每坪kW", "設計負荷(kW)", "設計負荷(kcal)", "機型", "數量", "單台能力(kW)"]
        for idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=idx).value = h
            
        for r_idx, room in enumerate(rooms_data, 2):
            ws.cell(row=r_idx, column=1).value = room.get("room_name")
            ws.cell(row=r_idx, column=2).value = room.get("area_m2")
            ws.cell(row=r_idx, column=3).value = room.get("ping_val")
            ws.cell(row=r_idx, column=4).value = room.get("final_suggested_kcal_per_ping")
            ws.cell(row=r_idx, column=5).value = room.get("kw_per_ping")
            ws.cell(row=r_idx, column=6).value = room.get("total_load_kw")
            ws.cell(row=r_idx, column=7).value = room.get("total_load_kcal")
            ws.cell(row=r_idx, column=8).value = room.get("indoor_model")
            ws.cell(row=r_idx, column=9).value = room.get("qty")
            ws.cell(row=r_idx, column=10).value = room.get("indoor_capacity_kw")
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
